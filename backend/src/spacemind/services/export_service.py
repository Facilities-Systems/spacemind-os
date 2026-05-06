"""
SpaceMind OS — Export Service
Converts a DecompositionResult into PDF, Markdown, or JSON.
Uses reportlab for PDF generation.
"""
from __future__ import annotations

import io
from datetime import UTC, datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from spacemind.domain.schemas import DecompositionResult

# ─── Brand colours (dark-mode brand palette mapped to print) ─────────────────
BRAND_BLUE = colors.HexColor("#6366F1")
BRAND_DARK = colors.HexColor("#1E293B")
BRAND_LIGHT = colors.HexColor("#F8FAFC")
ACCENT_AMBER = colors.HexColor("#F59E0B")
RISK_RED = colors.HexColor("#EF4444")
RISK_YELLOW = colors.HexColor("#EAB308")
RISK_GREEN = colors.HexColor("#22C55E")

_RISK_COLOUR = {"high": RISK_RED, "medium": RISK_YELLOW, "low": RISK_GREEN}


class ExportService:
    # ─── PDF ─────────────────────────────────────────────────────────────────

    def to_pdf(self, result: DecompositionResult) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        story = []

        # ── Title block ──────────────────────────────────────────────────────
        title_style = ParagraphStyle(
            "SMTitle",
            parent=styles["Title"],
            textColor=BRAND_BLUE,
            fontSize=20,
            spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            "SMSub",
            parent=styles["Normal"],
            textColor=colors.grey,
            fontSize=9,
            spaceAfter=12,
        )
        body_style = ParagraphStyle(
            "SMBody",
            parent=styles["Normal"],
            fontSize=9,
            leading=13,
            spaceAfter=6,
        )
        h2_style = ParagraphStyle(
            "SMH2",
            parent=styles["Heading2"],
            textColor=BRAND_DARK,
            fontSize=12,
            spaceBefore=14,
            spaceAfter=4,
        )
        h3_style = ParagraphStyle(
            "SMH3",
            parent=styles["Heading3"],
            textColor=BRAND_BLUE,
            fontSize=10,
            spaceBefore=8,
            spaceAfter=2,
        )

        story.append(Paragraph("SpaceMind OS — Execution Plan", title_style))
        story.append(Paragraph(
            f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')} &nbsp;|&nbsp; "
            f"ID: {result.id[:8]}",
            sub_style,
        ))
        story.append(HRFlowable(width="100%", thickness=1, color=BRAND_BLUE))
        story.append(Spacer(1, 0.3 * cm))

        # ── Summary box ──────────────────────────────────────────────────────
        summary_data = [
            ["Request", result.request_summary],
            ["Type", result.request_type.value.replace("_", " ").title()],
            ["Location", result.location_context.location_id],
            ["Priority", result.priority.upper()],
            ["Est. Duration", f"{result.total_estimated_duration_days or '—'} days"],
            ["Total Tasks", str(result.total_tasks)],
        ]
        summary_table = Table(summary_data, colWidths=[3.5 * cm, 13 * cm])
        summary_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("BACKGROUND", (0, 0), (0, -1), BRAND_DARK),
            ("BACKGROUND", (1, 0), (1, -1), BRAND_LIGHT),
            ("ROWBACKGROUNDS", (1, 0), (1, -1), [BRAND_LIGHT, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.4 * cm))

        # ── Phases & tasks ───────────────────────────────────────────────────
        story.append(Paragraph("Execution Phases", h2_style))
        for phase in result.phases:
            story.append(Paragraph(
                f"Phase {phase.order}: {phase.name}", h3_style,
            ))
            if phase.description:
                story.append(Paragraph(phase.description, body_style))

            if phase.tasks:
                task_data = [["#", "Task", "Owner", "Est. Hours", "Risk"]]
                for i, task in enumerate(phase.tasks, 1):
                    task_data.append([
                        str(i),
                        task.name,
                        task.responsible.party.value.replace("_", " ").title(),
                        str(task.estimated_duration_hours or "—"),
                        task.risk_level.value.title(),
                    ])
                task_table = Table(
                    task_data,
                    colWidths=[0.7 * cm, 7.5 * cm, 3.5 * cm, 2 * cm, 2.3 * cm],
                )
                task_style = TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ])
                # Colour risk column
                for row_idx, task in enumerate(phase.tasks, 1):
                    rc = _RISK_COLOUR.get(task.risk_level.value, RISK_GREEN)
                    task_style.add("TEXTCOLOR", (4, row_idx), (4, row_idx), rc)
                task_table.setStyle(task_style)
                story.append(task_table)
            story.append(Spacer(1, 0.2 * cm))

        # ── Key risks ────────────────────────────────────────────────────────
        if result.key_risks:
            story.append(Paragraph("Key Risks", h2_style))
            for risk in result.key_risks:
                story.append(Paragraph(f"• {risk}", body_style))

        # ── Recommendations ──────────────────────────────────────────────────
        if result.recommendations:
            story.append(Paragraph("Recommendations", h2_style))
            for rec in result.recommendations:
                story.append(Paragraph(f"• {rec}", body_style))

        # ── Landlord items ───────────────────────────────────────────────────
        if result.landlord_items:
            story.append(Paragraph("Landlord Approval Required", h2_style))
            for item in result.landlord_items:
                story.append(Paragraph(f"⚠ {item}", body_style))

        # ── Footer ───────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
        story.append(Paragraph(
            "SpaceMind OS — AI-powered Facilities Operations Intelligence",
            ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, textColor=colors.grey),
        ))

        doc.build(story)
        return buffer.getvalue()

    # ─── Excel ────────────────────────────────────────────────────────────────

    def to_excel(self, result: DecompositionResult) -> bytes:
        """
        Export the decomposition plan as a multi-sheet Excel workbook.
        Adopted from UFM project pattern: facilities teams live in Excel.

        Sheets:
          1. Summary   — plan metadata
          2. Tasks     — all tasks flat (one row per task, phase column)
          3. Risks     — key risks + per-task risks
          4. Compliance — compliance notes
        """
        import io
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # ── Build data frames ────────────────────────────────────────────────
        summary_data = {
            "Field": [
                "Plan ID", "Generated", "Request Summary", "Request Type",
                "Location", "Country", "Priority",
                "Total Phases", "Total Tasks", "Est. Duration",
            ],
            "Value": [
                result.id[:8],
                datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"),
                result.request_summary,
                result.request_type.value.replace("_", " ").title(),
                result.location_context.location_id,
                result.location_context.country,
                result.priority.upper(),
                len(result.phases),
                result.total_tasks,
                f"{result.total_estimated_duration_days or '—'} days",
            ],
        }

        tasks_rows = []
        for phase in sorted(result.phases, key=lambda p: p.order):
            for task in phase.tasks:
                tasks_rows.append({
                    "Phase #":       phase.order,
                    "Phase":         phase.name,
                    "Task":          task.name,
                    "Description":   task.description or "",
                    "Owner":         task.responsible.party.value.replace("_", " ").title(),
                    "Est. Hours":    task.estimated_duration_hours or "",
                    "Risk Level":    task.risk_level.value.title(),
                    "Status":        task.status.value.replace("_", " ").title(),
                    "Landlord":      "Yes" if task.landlord_approval_required else "No",
                    "Notes":         task.notes or "",
                })

        risks_rows = []
        for r in result.key_risks:
            risks_rows.append({"Level": "Project", "Risk": r, "Task": ""})
        for phase in result.phases:
            for task in phase.tasks:
                for r in task.risks:
                    risks_rows.append({"Level": task.risk_level.value.title(), "Risk": r, "Task": task.name})

        compliance_rows = [{"Note": n} for n in result.compliance_notes]
        recommendation_rows = [{"Recommendation": r} for r in result.recommendations]

        # ── Write to Excel ───────────────────────────────────────────────────
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
            pd.DataFrame(tasks_rows).to_excel(writer, sheet_name="Tasks", index=False)

            if risks_rows:
                pd.DataFrame(risks_rows).to_excel(writer, sheet_name="Risks", index=False)
            if compliance_rows:
                pd.DataFrame(compliance_rows).to_excel(writer, sheet_name="Compliance", index=False)
            if recommendation_rows:
                pd.DataFrame(recommendation_rows).to_excel(writer, sheet_name="Recommendations", index=False)

            # ── Styling ──────────────────────────────────────────────────────
            wb = writer.book
            header_fill = PatternFill("solid", fgColor="1E293B")
            header_font = Font(bold=True, color="FFFFFF")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Style header row
                for cell in ws[1]:
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = Alignment(horizontal="left")
                # Auto-fit columns (approximate)
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value or "")) for cell in col), default=10
                    )
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        return buffer.getvalue()

    # ─── Markdown ─────────────────────────────────────────────────────────────

    def to_markdown(self, result: DecompositionResult) -> str:
        lines: list[str] = []
        lines.append(f"# SpaceMind OS — Execution Plan")
        lines.append(f"")
        lines.append(f"**Generated:** {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}  ")
        lines.append(f"**Plan ID:** `{result.id[:8]}`")
        lines.append(f"")
        lines.append(f"---")
        lines.append(f"")
        lines.append(f"## Summary")
        lines.append(f"")
        lines.append(f"| Field | Value |")
        lines.append(f"|-------|-------|")
        lines.append(f"| Request | {result.request_summary} |")
        lines.append(f"| Type | {result.request_type.value.replace('_', ' ').title()} |")
        lines.append(f"| Location | {result.location_context.location_id} |")
        lines.append(f"| Priority | {result.priority.upper()} |")
        lines.append(f"| Est. Duration | {result.total_estimated_duration_days or '—'} days |")
        lines.append(f"| Total Tasks | {result.total_tasks} |")
        lines.append(f"")

        lines.append(f"## Execution Phases")
        lines.append(f"")
        for phase in result.phases:
            lines.append(f"### Phase {phase.order}: {phase.name}")
            if phase.description:
                lines.append(f"")
                lines.append(f"{phase.description}")
            lines.append(f"")
            if phase.tasks:
                lines.append(f"| # | Task | Owner | Est. Hours | Risk |")
                lines.append(f"|---|------|-------|------------|------|")
                for i, task in enumerate(phase.tasks, 1):
                    owner = task.responsible.party.value.replace("_", " ").title()
                    lines.append(
                        f"| {i} | {task.name} | {owner} "
                        f"| {task.estimated_duration_hours or '—'} "
                        f"| {task.risk_level.value.title()} |"
                    )
                lines.append(f"")

        if result.key_risks:
            lines.append(f"## Key Risks")
            lines.append(f"")
            for risk in result.key_risks:
                lines.append(f"- {risk}")
            lines.append(f"")

        if result.recommendations:
            lines.append(f"## Recommendations")
            lines.append(f"")
            for rec in result.recommendations:
                lines.append(f"- {rec}")
            lines.append(f"")

        if result.landlord_items:
            lines.append(f"## Landlord Approval Required")
            lines.append(f"")
            for item in result.landlord_items:
                lines.append(f"- ⚠ {item}")
            lines.append(f"")

        lines.append(f"---")
        lines.append(f"*SpaceMind OS — AI-powered Facilities Operations Intelligence*")
        return "\n".join(lines)
